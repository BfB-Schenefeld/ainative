#!/usr/bin/env python3
"""
export_quiz.py -- export quiz records to a human-readable workbook and CSV.

Reads knowledge/courses/<slug>/quiz/*.yaml and writes:
  <out>.xlsx   two sheets: one row per question (wide), one row per option (long)
  <out>.csv    semicolon-delimited, UTF-8 BOM, for German Excel

Usage:
    python3 tools/export_quiz.py --course ai-native-safe-overview --out exports/final-quiz
"""
from __future__ import annotations

import argparse
import csv
import pathlib
import sys

import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).resolve().parent.parent

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BASE_FONT = Font(name="Arial", size=10)
BOLD = Font(name="Arial", size=10, bold=True)
OK_FILL = PatternFill("solid", fgColor="D9EAD3")
NO_FILL = PatternFill("solid", fgColor="FCE8E6")
ALT_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOPWRAP = Alignment(wrap_text=True, vertical="top")


def load(course: str):
    qdir = ROOT / "knowledge" / "courses" / course / "quiz"
    items = []
    for f in sorted(qdir.glob("*.yaml")):
        items.append(yaml.safe_load(f.read_text(encoding="utf-8")))
    return items


def load_lesson_titles(course: str) -> dict:
    """Real lesson titles from course.yaml, so the export shows names not slugs."""
    cfile = ROOT / "knowledge" / "courses" / course / "course.yaml"
    try:
        doc = yaml.safe_load(cfile.read_text(encoding="utf-8"))
        return {les["id"]: les["title"] for les in doc.get("lessons", []) or []}
    except Exception:
        return {}


LESSON_TITLES: dict = {}


def lesson_label(item) -> str:
    ref = item.get("lesson") or ""
    if ref in LESSON_TITLES:
        return LESSON_TITLES[ref]
    tail = ref.rsplit("/", 1)[-1] if ref else ""
    return tail.replace("-", " ").strip() if tail else "unbekannt"


def wide_rows(items, maxopt):
    """One row per question: every option, its verdict, and its explanation."""
    header = ["Nr", "ID", "Typ", "Schwierigkeit", "Lektion", "Frage"]
    for i in range(maxopt):
        L = chr(65 + i)
        header += [f"Option {L}", f"{L} richtig?", f"Begründung {L}"]
    header += ["Richtige Antwort", "Begründung (richtig)", "Erläuterung zur Frage",
               "Antwortstatus", "Quelle"]

    rows = []
    for n, it in enumerate(items, 1):
        opts = it.get("options", []) or []
        correct = next((o for o in opts if o.get("correct")), None)
        row = [
            n,
            it.get("id", ""),
            "Einfachauswahl" if it.get("type") == "single-choice" else it.get("type", ""),
            {"easy": "leicht", "medium": "mittel", "harder": "schwer"}.get(
                it.get("difficulty", ""), it.get("difficulty", "")),
            lesson_label(it),
            it.get("stem", ""),
        ]
        for i in range(maxopt):
            if i < len(opts):
                o = opts[i]
                row += [o.get("text", ""), "RICHTIG" if o.get("correct") else "falsch", o.get("feedback", "")]
            else:
                row += ["", "", ""]
        row += [
            correct.get("text", "") if correct else "",
            correct.get("feedback", "") if correct else "",
            it.get("rationale", ""),
            it.get("answer_status", ""),
            (it.get("sources") or [""])[0],
        ]
        rows.append(row)
    return header, rows


def long_rows(items):
    """One row per option: easier to filter, sort and pivot."""
    header = ["Nr", "Frage-ID", "Schwierigkeit", "Lektion", "Frage", "Option",
              "Antworttext", "Richtig?", "Begründung", "Erläuterung zur Frage"]
    rows = []
    for n, it in enumerate(items, 1):
        for o in it.get("options", []) or []:
            rows.append([
                n,
                it.get("id", ""),
                {"easy": "leicht", "medium": "mittel", "harder": "schwer"}.get(
                    it.get("difficulty", ""), it.get("difficulty", "")),
                lesson_label(it),
                it.get("stem", ""),
                (o.get("key") or "").upper(),
                o.get("text", ""),
                "RICHTIG" if o.get("correct") else "falsch",
                o.get("feedback", ""),
                it.get("rationale", "") if o.get("correct") else "",
            ])
    return header, rows


def style_sheet(ws, header, rows, widths, wrap_cols, verdict_cols, row_height=None):
    ws.append(header)
    for r in rows:
        ws.append(r)
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    for cell in ws[1]:
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BOX
    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"
    for ri, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.font = BASE_FONT
            cell.border = BOX
            if cell.column in wrap_cols:
                cell.alignment = TOPWRAP
            else:
                cell.alignment = Alignment(vertical="top")
            if cell.column in verdict_cols:
                cell.alignment = Alignment(vertical="top", horizontal="center")
                if cell.value == "RICHTIG":
                    cell.fill = OK_FILL
                    cell.font = BOLD
                elif cell.value == "falsch":
                    cell.fill = NO_FILL
            elif ri % 2 == 0 and cell.column not in verdict_cols:
                cell.fill = ALT_FILL
        if row_height:
            ws.row_dimensions[ri].height = row_height


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--course", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args()

    global LESSON_TITLES
    items = load(args.course)
    if not items:
        print("no quiz items found", file=sys.stderr)
        return 1
    LESSON_TITLES = load_lesson_titles(args.course)
    # only emit as many option columns as the bank actually uses
    maxopt = max((len(it.get("options") or []) for it in items), default=0)

    out = pathlib.Path(args.out)
    out.parent.mkdir(parents=True, exist_ok=True)

    wh, wr = wide_rows(items, maxopt)
    lh, lr = long_rows(items)

    wb = Workbook()
    ws = wb.active
    ws.title = "Fragen"
    FIRST = 7  # first option column after Nr, ID, Typ, Schwierigkeit, Lektion, Frage
    wide_widths = [5, 40, 15, 14, 40, 60]
    for _ in range(maxopt):
        wide_widths += [50, 12, 62]
    wide_widths += [50, 62, 66, 14, 42]
    last = FIRST + maxopt * 3
    wide_wrap = ({5, 6} | {c for c in range(FIRST, last) if (c - FIRST) % 3 != 1}
                 | {last, last + 1, last + 2})
    wide_verdict = {c for c in range(FIRST, last) if (c - FIRST) % 3 == 1}
    style_sheet(ws, wh, wr, wide_widths, wide_wrap, wide_verdict, row_height=110)

    ws2 = wb.create_sheet("Antwortoptionen")
    style_sheet(ws2, lh, lr, [5, 40, 14, 40, 60, 8, 55, 12, 70, 66],
                {4, 5, 7, 9, 10}, {6, 8}, row_height=58)

    info = wb.create_sheet("Info")
    for r in [
        ["AI-Native SAFe Overview — Final Quiz", ""],
        ["", ""],
        ["Fragen im Pool", len(items)],
        ["Antwortoptionen je Frage", maxopt],
        ["Schwierigkeitsverteilung", ""],
        ["Fragen pro Versuch", 15],
        ["Bestehensgrenze", "12 von 15 (80 %)"],
        ["Modus", "Open Book, unbegrenzte Wiederholungen, ohne Zeitlimit"],
        ["", ""],
        ["Blatt „Fragen“", "Eine Zeile je Frage, alle Optionen nebeneinander."],
        ["Blatt „Antwortoptionen“", "Eine Zeile je Antwortoption — zum Filtern und Sortieren."],
        ["", ""],
        ["Quelle", "src:2026-08-26/upgrade-path-final-quiz"],
        ["Herkunft", "Bundle-Chunk FinalQuiz-DUvUYCnK.js der Upgrade-Path-Webanwendung"],
        ["Vorbehalt", "Das Original-Artefakt wurde nach dem Auslesen neu deployt; "
                      "die Transkription ist nicht byte-verifiziert. Siehe source.yaml."],
        ["Rechte", "Scaled Agile, Inc. — keine Weitergabe (rights.redistribution: none)"],
    ]:
        info.append(r)
    bands = {}
    for it in items:
        d = it.get("difficulty")
        if d:
            bands[d] = bands.get(d, 0) + 1
    if bands:
        label = {"easy": "leicht", "medium": "mittel", "harder": "schwer"}
        info["B6"] = ", ".join(f"{label.get(k, k)}: {v}" for k, v in
                               sorted(bands.items(), key=lambda kv: ["easy", "medium", "harder"].index(kv[0])
                                      if kv[0] in ("easy", "medium", "harder") else 9))
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 95
    for row in info.iter_rows():
        for cell in row:
            cell.font = BASE_FONT
            cell.alignment = TOPWRAP
    info["A1"].font = Font(name="Arial", size=13, bold=True)
    for r in (3, 4, 5, 6, 8, 9, 11, 12, 13, 14):
        info[f"A{r}"].font = BOLD

    xlsx = out.with_suffix(".xlsx")
    wb.save(xlsx)

    csv_path = out.with_suffix(".csv")
    with csv_path.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";", quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        w.writerow(wh)
        for r in wr:
            w.writerow([str(c).replace("\r\n", " ").replace("\n", " ") if c is not None else "" for c in r])

    csv_long = out.parent / (out.name + "-optionen.csv")
    with csv_long.with_suffix(".csv").open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh, delimiter=";", quoting=csv.QUOTE_ALL, lineterminator="\r\n")
        w.writerow(lh)
        for r in lr:
            w.writerow([str(c).replace("\n", " ") if c is not None else "" for c in r])

    print(f"{len(items)} questions, {len(lr)} options -> {xlsx.name}, {csv_path.name}, "
          f"{csv_long.with_suffix('.csv').name}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
